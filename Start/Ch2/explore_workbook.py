# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Open, load, and explore workbook content 

import openpyxl

filename = "FinancialSample.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(filename)

# Print basic information
print('# worksheets: ', len(wb.sheetnames))
for worksheet_name in wb.sheetnames:
  worksheet = wb[worksheet_name]
  print(f"\nWorksheet: {worksheet_name}")
  # Explore each worksheet
  # # Get dimensions
  dimensions = worksheet.dimensions
  print(f'      - Dimensions: {dimensions}')

  print(f'Min row: {worksheet.min_row}')
  print(f'Max row: {worksheet.max_row}')

  print(f'Min col: {worksheet.min_column}')
  print(f'Max col: {worksheet.max_column}')
  
  # Check if the worksheet is empty
  if worksheet.max_row==1 and worksheet.max_column==1:
    print("     - Worksheet is empty")
  else:
    cell = worksheet["A1"]
    print(f"    - Top-left cell value: {cell.value}")
    cell = worksheet.cell(row=worksheet.max_row,column=worksheet.max_column)
    print(f"    - Bottom-right cell value: {cell.value}")
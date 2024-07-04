# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Apply conditional formatting to a worksheet 

import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle


filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)
sheet = workbook["SalesData"]

# define the style to represent the formatting
red_color = "ffd2d2"
bold_text = Font(bold=True, color="00ff0000")
red_fill = PatternFill(bgColor=red_color, fill_type="solid")

# Conditional formatting requires DifferentialStyle from openpyxl.styles.differential
diff_style = DifferentialStyle(font=bold_text, fill=red_fill)

# create a rule for the condition (uses openpyxl.formatting.Rule)
rule = Rule(type="expression", dxf=diff_style, formula=["$L1<10000"]) # rule is for cell L1
# "$L1<10000" --> we set a specific column with $, notice row 1 doesn't have $ here. 
# this will allow the rule iterate over the rows.
# if we use $L2 in the rule: use next row value on the current row's format
# ($L0 leads to excell error); the row number in the rule is the starting value for 
# the iteration. 

# add the rule to the entire sheet (i.e. entire row not just cell L1)
# we want to apply rule on col L
dimensions = "L1:L701"# to format entire sheet (row by row) use: sheet.dimensions
sheet.conditional_formatting.add(dimensions, rule) 

workbook.save("CondFormat.xlsx")
print("Workbook created successfully!")
